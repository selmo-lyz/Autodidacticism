# 注意事項
# 圖檔要放在 ./images 底下

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Alignment, Font, Border, Side
from openpyxl.drawing.image import Image
from os.path import isfile

# 以額外的 xlsx 檔，依照指定欄位擴充原資料
def extendDF(input_df, extend_df, origin_cols, extend_cols):
    # 在原資料中增加額外欄位
    for i in range(0, len(extend_cols)):
        input_df[extend_cols[i]] = [0 for x in range(0, input_df.shape[0])]
    origin_cols = origin_cols + extend_cols

    # 以學號識別該資料是否為新生，並在原資料中添加新生資料且刪除非新生者
    input_df.set_index(cols[0], inplace=True)
    extend_df.set_index("學號", inplace=True)

    for i in input_df.index:
        exist = False
        for j in extend_df.index:
            print(j)
            if i == j:
                exist = True
                for k in extend_cols:
                    input_df.loc[i, k] = extend_df.loc[j, k]
                extend_df = extend_df.drop(index=j)
                break;
        if not exist:
            input_df = input_df.drop(index=i)

    input_df.reset_index(inplace=True)

    return (input_df, origin_cols)
    
# 將原資料轉換為指定格式
def variant_transpose(input_df, origin_cols, transpose_rows, col2row_dict, num_cols_in_row = 6):
    output_df = pd.DataFrame()
    temp_df = pd.DataFrame()

    for i in input_df.index:
        col = i % num_cols_in_row
        if col == 0:
            if i == num_cols_in_row:
                output_df = temp_df
            else:
                output_df = pd.concat([output_df, temp_df])
            temp_df = pd.DataFrame(index=transpose_rows, columns=[i for i in range(0, num_cols_in_row)])

        for j in input_df.columns:
            temp_df.loc[col2row_dict[j], col] = input_df.loc[i, j]
    
    return output_df

# 將 pandas 的 Dataframe 轉為 openpyxl 的 Workbook
def loadDF(input_df):
    # 以 openpyxl 的格式使用 output_df
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(input_df, index=True, header=False):
        ws.append(r)

    return wb

# 設定表格格式
def setStyle(input_wb, rows):
    ws = input_wb.active
    # 全儲存格格式
    all_cell = NamedStyle(name='all_cell')
    all_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    all_cell.font = Font(name='Calibri', size=11)
    # 儲存格邊框格式
    #border_l = Border(left=Side(border_style='thin'))
    #border_r = Border(right=Side(border_style='thin'))
    border_t = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'))
    border_b = Border(top=Side(border_style='thin'), bottom=Side(border_style='medium'))
    border_m = Border(bottom=Side(border_style='thin'))
    # 各標籤於 row 的位置
    row_pos_pic = 0
    row_pos_addr = 0

    for r in range(1, len(rows) + 1):
        if ws.cell(r, 1).value == "照片":
            row_pos_pic = r
        if ws.cell(r, 1).value == "住址":
            row_pos_addr = r

    # 設定格式：全儲存格
    for r in range(1, ws.max_row):
        ws.row_dimensions[r].height = 30
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).style = all_cell

    # 設定格式：欄
    for c in range(66, 66 + ws.max_column + 1):
        ws.column_dimensions[chr(c)].width = 23 

    # 設定格式：標籤欄
    ws.column_dimensions['A'].width = 11

    # 設定格式：照片
    for r in range(row_pos_pic, ws.max_row,11):
        ws.row_dimensions[r].height = 160

    # 設定格式：住址
    for r in range(row_pos_addr, ws.max_row,11):
        ws.row_dimensions[r].height = 50

    # 設定格式：border
    for r in range(2, ws.max_row):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = border_m

    for r in range(2, ws.max_row, 11):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = border_t
            ws.cell(r + 9, c).border = border_b
    
    return input_wb

# 載入圖片至指定學號的照片欄
def loadImg(input_wb, rows):
    ws = input_wb.active

    # magni: 圖片放大倍率
    magni = 1.5;
    # 各標籤於 row 的位置
    row_pos_pic = 0
    row_pos_ssn = 0

    for r in range(1, len(rows) + 1):
        if ws.cell(r, 1).value == "照片":
            row_pos_pic = r
        if ws.cell(r, 1).value == "學號":
            row_pos_ssn = r

    # 插入相片
    for r in range(0, ws.max_row, len(rows)):
        for c in range(2, ws.max_column + 1):
            file_name = "./images/" + str(ws.cell(r + row_pos_ssn, c).value) + ".jpg"
            if (isfile(file_name)):
                img = Image(file_name)
                img.width = 106 * magni # pixel
                img.height = 132.3 * magni # pixel
                ws.add_image(img, ws.cell(r + row_pos_pic, c).coordinate)
    
    return input_wb

# 變數說明
# input_df: 儲存讀取進來的 xlsx 資料
# input_df_fresher: 儲存讀取進來的 xlsx 資料 part 2
# output_name: 輸出檔案名稱
# cols: 原資料中所需的欄位 (column/header) 名稱
# extend_cols: 追加資料中所需的欄位名稱
# rows: 新資料中所需的列 (row/index) 名稱
# col2row_dict: "原資料欄位" 與 "其所對應列" 的集合
input_df = pd.read_excel('input.xlsx')
extend_df_fresher = pd.read_excel('fresher.xlsx')
output_name = "output.xlsx"
cols = [' 學號', ' 姓名', ' 地址', ' 聯絡電話1', ' 聯絡電話2', ' 導師姓名']
extend_cols = ['出生年月日', '入學前畢業學校']
rows = ['學號', '照片', '中文姓名', '英文姓名', '出生年月日', '住址', '住家電話', '手機號碼', '指導教授', '畢業學校', '']
col2row_dict = {' 學號': '學號', ' 姓名': '中文姓名', '出生年月日': '出生年月日', ' 地址': '住址', ' 聯絡電話1': '住家電話', ' 聯絡電話2': '手機號碼', ' 導師姓名': '指導教授', '入學前畢業學校': '畢業學校'}
num_cols_in_row = 6


# --------------------
# 轉換表格排列格式
# --------------------

# 過濾原資料中無用資料欄
input_df = input_df[cols]

# 追加額外資料進原資料中
(input_df, cols) = extendDF(input_df, extend_df_fresher, cols, extend_cols)
output_df = variant_transpose(input_df, cols, rows, col2row_dict, num_cols_in_row)

# Debug
#print(output_df)
#print(output_df.shape)

# --------------------
# 設定格式
# --------------------

wb = loadDF(output_df)
wb = setStyle(wb, rows)
wb = loadImg(wb, rows)

# 儲存表格
wb.save(output_name)